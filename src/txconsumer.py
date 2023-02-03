import datetime
from dataclasses import dataclass
from enum import Enum
from functools import partial
from typing import Any, Iterator, Tuple, Callable, Optional

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import numbers

from tx import Transaction, Category


class TransactionWorkbookWriter:
    class Column(Enum):
        charge = 'a', 10, "סכום החיוב"
        business = 'b', 30, "בית עסק"
        transaction_date = 'c', 20, "תאריך עסקה"
        category = 'd', 13, "טיב"
        details = 'e', 20, "פירוט"
        card = 'f', 10, "כרטיס"
        notes = 'g', 20, "הערות"
        transaction_sum = 'h', 15, "סכום העסקה"

        def __init__(self, position: str, width: int, description: str) -> None:
            self.position = position
            self.width = width
            self.description = description

        @staticmethod
        def by_position():
            return {col.position: col.description for col in TransactionWorkbookWriter.Column}

        @staticmethod
        def width_per_column():
            return {col.position: col.width for col in TransactionWorkbookWriter.Column}

    @dataclass
    class Filter:
        begin: Optional[datetime.datetime] = None
        end: Optional[datetime.datetime] = None
        excludebusiness: Tuple[str] = ()

    def __init__(self, outfile: str, txfilter: Filter) -> None:
        self._wb = Workbook()
        self._sheet = self._wb.active
        self._sheet.append(TransactionWorkbookWriter.Column.by_position())
        self._sheet.sheet_view.rightToLeft = True
        for column, width in TransactionWorkbookWriter.Column.width_per_column().items():
            self._sheet.column_dimensions[column].width = width
        self._outfile = outfile
        self._filter = txfilter if txfilter else TransactionWorkbookWriter.Filter()

    def __enter__(self) -> Any:
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self._sheet.freeze_panes = self._sheet["a2"]
        self._add_sort_dropdown()
        self._add_summary()
        self._wb.save(self._outfile)

    def _add_sort_dropdown(self) -> None:
        last_col = chr(ord('a') - 1 + self._sheet.max_column)
        last_row = self._sheet.max_row
        self._sheet.auto_filter.ref = f"a1:{last_col}{last_row}"
        self._sheet.auto_filter.add_sort_condition(f"a2:a{last_row}")

    def accept(self, transactions: Iterator[Transaction]) -> None:
        for transaction in transactions:
            if self._relevant(transaction):
                self._sheet.append(self._convert(transaction))
                self._set_number_format(TransactionWorkbookWriter.Column.charge.position)

    def _set_number_format(self, position: str) -> None:
        self._sheet[f"{position}{self._sheet.max_row}"].number_format = numbers.FORMAT_NUMBER

    @staticmethod
    def _convert(transaction: Transaction) -> Tuple:
        return (
            transaction.amount,
            transaction.business,
            transaction.transaction_date,
            transaction.category.value,
            transaction.details,
            transaction.card,
            transaction.notes,
            transaction.transaction_sum,
        )

    def _relevant(self, transaction: Transaction) -> bool:
        for business in self._filter.excludebusiness:
            if business in transaction.business:
                return False
        if self._filter.begin and self._filter.begin > transaction.charge_date:
            return False
        if self._filter.end and self._filter.end < transaction.charge_date:
            return False
        return True

    def _add_summary_row(self, description: str, formula: str) -> None:
        self._sheet.append((description, formula))
        self._set_number_format(position='b')

    def _add_category(self, cat: Category, charge_range: str, category_range: str) -> None:
        self._add_summary_row(
            description=cat.value,
            formula=f"=SUMIFS({charge_range}, {category_range}, \"{cat.value}\")",
        )

    def _add_summary(self) -> None:
        last_data_row = self._sheet.max_row
        charge_pos = TransactionWorkbookWriter.Column.charge.position
        cat_pos = TransactionWorkbookWriter.Column.category.position
        charge_range = f"{charge_pos}2:{charge_pos}{last_data_row}"
        category_range = f"{cat_pos}2:{cat_pos}{last_data_row}"
        add_category = partial(self._add_category, charge_range=charge_range, category_range=category_range)
        gap = 3
        self._add_rows(gap)
        self._add_category_summary(add_category, start_row=last_data_row + gap + 1)
        self._add_rows(1)
        self._add_totals_summary(add_category, charge_range)

    def _add_totals_summary(self, add_category: Callable, charge_range: str) -> None:
        self._add_summary_row(
            description="הוצאות",
            formula=f"=SUMIFS({charge_range}, {charge_range}, \">0\")",
        )
        add_category(Category.income)
        self._add_summary_row(
            description="סך הוצאות",
            formula=f"=SUM({charge_range})",
        )

    def _add_category_summary(self, add_category: Callable, start_row: int) -> None:
        for category in Category:
            if category is not Category.income:
                add_category(category)
        self._add_category_chart(start_row=start_row, end_row=self._sheet.max_row, start_col=2, end_col=2)

    def _add_rows(self, rows: int) -> None:
        for _ in range(rows):
            self._sheet.append(())

    def _add_category_chart(self, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        chart = PieChart()
        data = Reference(worksheet=self._sheet, min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col)
        chart.add_data(data, from_rows=False, titles_from_data=False)
        cats = Reference(self._sheet, min_col=1, min_row=start_row, max_col=1, max_row=end_row)
        chart.set_categories(cats)
        self._sheet.add_chart(chart, f'c{start_row}')
