import datetime
from abc import ABC, abstractmethod
from dataclasses import dataclass
from enum import Enum
from functools import partial
from typing import Any, Iterator, Tuple, Dict, Set, Callable, Iterable, TypeVar

from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import numbers


class Category(Enum):
    mortgage = "משכנתא"
    food = "אוכל"
    education = "חינוך"
    running_expenses = "שוטף"
    mentoring = "הדרכה"
    donation = "תרומה"
    tax = "מס"
    insurance = "ביטוח"
    atm = "כספומט"
    fuel = "דלק"
    savings = "חסכון"
    transport = "תחבורה"
    other = "אחר"
    income = "הכנסות"


descriptions_by_category: Dict[Category, Set[str]] = {
    Category.mortgage: {
        "משכנתא",
    },
    Category.tax: {
        "מסים",
    },
    Category.running_expenses: {
        "ועד",
        "אינטרנט",
        "חברת החשמל",
        "019",
        "פלאפון",
        "בזק",
    },
    Category.savings: {
        "חסכון",
    },
    Category.donation: {
        "פעמונים",
        "מוסדות חב\"ד",
        "מכון מאיר",
        "עטרת",
        "מה יפו פעמי",
        "התורה והארץ",
        "גרעין יפו",
        "בית דוד בית שמש",
        "המרכז העולמי לחסד",
    },
    Category.insurance: {
        "מכבי",
        "שירותי ברי",
        "ביטוח",
        "פניקס",
        "מגדל",
    },
    Category.education: {
        "אמונה",
    },
    Category.atm: {
        "כספומט"
    },
    Category.mentoring: {
        "שר שלום",
    },
    Category.fuel: {
        "פנגו",
        "פז",
        "כלל חובה",
    },
    Category.food: {
        "מכולת",
        "יינות ביתן",
        "שופרסל",
        "רמי לוי",
    },
}

category_by_description: Dict[str, Category] = {
    keyword: cat for cat, keywords in descriptions_by_category.items() for keyword in keywords
}


@dataclass
class Transaction:
    amount: Any
    business: str
    charge_date: datetime.datetime
    transaction_date: datetime.datetime
    details: str
    card: str
    notes: str
    transaction_sum: Any

    def __post_init__(self) -> None:
        self.category = self._compute_category()

    def _compute_category(self) -> Category:
        for kw, category in category_by_description.items():
            if kw in self.business:
                return category
        if self.amount < 0:
            return Category.income
        else:
            return Category.other


class TransactionProducer(ABC):

    def __init__(self, filename: str) -> None:
        workbook = load_workbook(filename=filename, read_only=True)
        self._row_gen = workbook.active.rows
        while self._is_header_row(next(self._row_gen)):
            pass

    @abstractmethod
    def _is_header_row(self, row) -> bool:
        """this method should return whether we have reached the row before the first data row"""

    def transaction_generator(self) -> Iterator[Transaction]:
        for row in self._row_gen:
            if not row[0].value:
                return
            yield self._convert(row)

    @abstractmethod
    def _convert(self, row) -> Transaction:
        """this method should convert a row to a Transaction"""


class TransactionWorkbookWriter:
    class Column(Enum):
        charge = 'a', 10, "סכום החיוב"
        business = 'b', 30, "בית עסק"
        transaction_date = 'c', 20, "תאריך עסקה"
        category = 'd', 13, "טיב"
        details = 'e', 20, "פירוט"
        card = 'f', 10, "כרטיס"
        notes = 'g', 20, "הערות"
        transaction_sum = 'h', 15, "סכום העסקה"

        def __init__(self, position: str, width: int, description: str) -> None:
            self.position = position
            self.width = width
            self.description = description

        @staticmethod
        def by_position():
            return {col.position: col.description for col in TransactionWorkbookWriter.Column}

        @staticmethod
        def width_per_column():
            return {col.position: col.width for col in TransactionWorkbookWriter.Column}

    def __init__(self, outfile: str, filters: dict) -> None:
        self._wb = Workbook()
        self._sheet = self._wb.active
        self._sheet.append(TransactionWorkbookWriter.Column.by_position())
        self._sheet.sheet_view.rightToLeft = True
        for column, width in TransactionWorkbookWriter.Column.width_per_column().items():
            self._sheet.column_dimensions[column].width = width
        self._outfile = outfile
        self._filters = filters

    def __enter__(self) -> Any:
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self._sheet.freeze_panes = self._sheet["a2"]
        self._add_sort_dropdown()
        self._add_summary()
        self._wb.save(self._outfile)

    def _add_sort_dropdown(self) -> None:
        last_col = chr(ord('a') - 1 + self._sheet.max_column)
        last_row = self._sheet.max_row
        self._sheet.auto_filter.ref = f"a1:{last_col}{last_row}"
        self._sheet.auto_filter.add_sort_condition(f"a2:a{last_row}")

    def accept(self, transactions: Iterator[Transaction]) -> None:
        for transaction in transactions:
            if self._relevant(transaction):
                self._sheet.append(self._convert(transaction))
                self._set_number_format(TransactionWorkbookWriter.Column.charge.position)

    def _set_number_format(self, position: str) -> None:
        self._sheet[f"{position}{self._sheet.max_row}"].number_format = numbers.FORMAT_NUMBER

    @staticmethod
    def _convert(transaction: Transaction) -> Tuple:
        return (
            transaction.amount,
            transaction.business,
            transaction.transaction_date,
            transaction.category.value,
            transaction.details,
            transaction.card,
            transaction.notes,
            transaction.transaction_sum,
        )

    def _relevant(self, transaction: Transaction) -> bool:
        return (
                self._filters and "month" in self._filters and transaction.charge_date.month == self._filters["month"]
                and
                "כרטיס ויזה" not in transaction.business
        )

    def _add_summary_row(self, description: str, formula: str) -> None:
        self._sheet.append((description, formula))
        self._set_number_format(position='b')

    def _add_category(self, cat: Category, charge_range: str, category_range: str) -> None:
        self._add_summary_row(
            description=cat.value,
            formula=f"=SUMIFS({charge_range}, {category_range}, \"{cat.value}\")",
        )

    def _add_summary(self) -> None:
        last_data_row = self._sheet.max_row
        charge_pos = TransactionWorkbookWriter.Column.charge.position
        cat_pos = TransactionWorkbookWriter.Column.category.position
        charge_range = f"{charge_pos}2:{charge_pos}{last_data_row}"
        category_range = f"{cat_pos}2:{cat_pos}{last_data_row}"
        add_category = partial(self._add_category, charge_range=charge_range, category_range=category_range)
        gap = 3
        self._add_rows(gap)
        self._add_category_summary(add_category, start_row=last_data_row + gap + 1)
        self._add_rows(1)
        self._add_totals_summary(add_category, charge_range)

    def _add_totals_summary(self, add_category: Callable, charge_range: str) -> None:
        self._add_summary_row(
            description="הוצאות",
            formula=f"=SUMIFS({charge_range}, {charge_range}, \">0\")",
        )
        add_category(Category.income)
        self._add_summary_row(
            description="סך הוצאות",
            formula=f"=SUM({charge_range})",
        )

    def _add_category_summary(self, add_category: Callable, start_row: int) -> None:
        for category in Category:
            if category is not Category.income:
                add_category(category)
        self._add_category_chart(start_row=start_row, end_row=self._sheet.max_row, start_col=2, end_col=2)

    def _add_rows(self, rows: int) -> None:
        for _ in range(rows):
            self._sheet.append(())

    def _add_category_chart(self, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        chart = PieChart()
        data = Reference(worksheet=self._sheet, min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col)
        chart.add_data(data, from_rows=False, titles_from_data=False)
        cats = Reference(self._sheet, min_col=1, min_row=start_row, max_col=1, max_row=end_row)
        chart.set_categories(cats)
        self._sheet.add_chart(chart, f'c{start_row}')


# todo: move to utils
T = TypeVar('T')


def series_combine_producers(producers: Iterable[T]) -> Iterable[T]:
    for prod in producers:
        for item in prod:
            yield item


class TransactionsMerger:

    def __init__(self, reports: Iterable[TransactionProducer]) -> None:
        self._reports = reports
        # self._combined_report = series_combine_producers(reports)

    def merge(self, transactions_processor: Any) -> None:
        with transactions_processor as processor:
            for report in self._reports:
                processor.accept(report.transaction_generator())
                # processor.accept(self._combined_report.transaction_generator())


class BankTransactions(TransactionProducer):
    def _is_header_row(self, row) -> bool:
        return row[0].value != "תאריך"

    def _convert(self, row) -> Transaction:
        return Transaction(
            amount=-row[3].value,
            business=row[2].value,
            transaction_date=row[0].value,
            charge_date=row[1].value,
            details="",
            card="",
            notes="",
            transaction_sum=None,
        )


class CreditTransactions(TransactionProducer):
    def _is_header_row(self, row) -> bool:
        return row[0].value != "כרטיס"

    def _convert(self, row) -> Transaction:
        return Transaction(
            amount=row[8].value,
            business=row[1].value,
            charge_date=datetime.datetime.strptime(row[7].value, "%d/%m/%Y"),
            transaction_date=datetime.datetime.strptime(row[2].value, "%d/%m/%Y"),
            details=row[6].value,
            card=row[0].value,
            notes="",
            transaction_sum=row[3].value,
        )


def main() -> None:
    bankfile = BankTransactions("/tmp/excel/ca.xlsx")
    creditfile = CreditTransactions("/tmp/excel/ashrai.xlsx")
    outfile = "/tmp/excel/merged.xlsx"
    filters = {"month": 6}
    merger = TransactionsMerger(reports=[bankfile, creditfile])
    merger.merge(TransactionWorkbookWriter(outfile=outfile, filters=filters))


if __name__ == "__main__":
    main()
